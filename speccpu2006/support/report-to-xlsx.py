#!/usr/bin/python3

import openpyxl
import os.path
import sys

from getopt import *
from openpyxl.styles import Font, Style, Color, colors, Alignment

# input variables
workbookname = 'unset'
is_android = False
is_llvm = False
is_pie = False
is_thumb = False
is_static = False
compiler_version = 'unset'
input_file = 'unset'
input_opt = 'unset'
os_name = 'linux'
arch_name = 'arm'
linkage = 'dynamic'
compiler_name = 'gcc'
# default 2 (for ARM/Thumb)
arch_mul = 2

# benchname, LLVM-compatible, Android-compatible, Diablo-compatible
benchmarks = (
  ['400.perlbench'  , True  , False , True  ],
  ['401.bzip2'      , True  , True  , True  ],
  ['403.gcc'        , True  , True  , True  ],
  ['410.bwaves'     , False , False , True  ],
  ['416.gamess'     , False , False , True  ],
  ['429.mcf'        , True  , True  , True  ],
  ['433.milc'       , True  , True  , True  ],
  ['434.zeusmp'     , False , False , True  ],
  ['435.gromacs'    , False , False , True  ],
  ['436.cactusADM'  , False , False , True  ],
  ['437.leslie3d'   , False , False , True  ],
  ['444.namd'       , True  , True  , True  ],
  ['445.gobmk'      , True  , True  , True  ],
  ['447.dealII'     , True  , True  , True  ],
  ['450.soplex'     , True  , True  , True  ],
  ['453.povray'     , True  , True  , False ],
  ['454.calculix'   , False , False , True  ],
  ['456.hmmer'      , True  , True  , True  ],
  ['458.sjeng'      , True  , True  , True  ],
  ['459.GemsFDTD'   , False , False , True  ],
  ['462.libquantum' , True  , False , True  ],
  ['464.h264ref'    , True  , True  , True  ],
  ['465.tonto'      , False , False , True  ],
  ['470.lbm'        , True  , True  , True  ],
  ['471.omnetpp'    , True  , True  , False ],
  ['473.astar'      , True  , False , True  ],
  ['481.wrf'        , False , False , True  ],
  ['482.sphinx_livepretend'    , True  , True  , True  ],
  ['483.Xalancbmk'  , True  , False , True  ],
  ['999.specrand'   , True  , True  , True  ],
)

opts = ('O0', 'O1', 'O2', 'O3', 'Os')
nr_opts = len(opts)

llvm_versions = ('3.2', '3.3', '3.4')
gcc_versions = ('4.6.4', '4.8.1')

os_names = ('linux', 'android')
arch_names = ('arm', 'thumb', 'i486')
linkage_names = ('static', 'dynamic', 'pie')
nr_linkages = len(linkage_names)
compiler_names = ('gcc', 'llvm')

header_col = 2
header_row = 1

benchname_col = 1
benchname_row = 5

compiler = 'unset'
system = 'unset'

centerStyle = Style(alignment=Alignment(horizontal='center', vertical='center'))

def parse_args():
  global workbookname, compiler_version, input_file, input_opt, os_name, linkage, compiler_name, arch_name

  short_opts = "f:v:i:o:O:l:c:a:"
  long_opts = ["output-file=","compiler-version=","input-file=","optlevel=","os=","linkage=","compiler=","arch="]

  try:
      opts, args = getopt(sys.argv[1:],short_opts,long_opts);
  except GetoptError:
      print("ERROR parsing options")
      sys.exit(-1)

  for opt, arg in opts:
    if opt == "-f" or opt == "--output-file":
      workbookname = arg
    elif opt == "-v" or opt == "--compiler-version":
      compiler_version = arg
    elif opt == "-i" or opt == "--input-file":
      input_file = arg
    elif opt == "-o" or opt == "--optlevel":
      input_opt = arg
    elif opt == "-O" or opt == "--os":
      os_name = arg
    elif opt == "-l" or opt == "--linkage":
      linkage = arg
    elif opt == "-c" or opt == "--compiler":
      compiler_name = arg
    elif opt == '-a' or opt == "--arch":
      arch_name = arg
    else:
      print("unknown option", opt)
      sys.exit(-1)

  if os_name not in os_names:
    print("invalid os name given, please use 'linux' or 'android'")
    sys.exit(-1)

  if arch_name not in arch_names:
    print("invalid architecture given, please use 'arm', 'thumb' or 'i486'")
    sys.exit(-1)

  if linkage not in linkage_names:
    print("invalid linkage given, please use 'static', 'dynamic' or 'pie'")
    sys.exit(-1)

  if compiler_name not in compiler_names:
    print("invalid compiler given, please use 'gcc' or 'llvm'")
    sys.exit(-1)

  if compiler_name == 'llvm' and (compiler_version not in llvm_versions):
    print("invalid version given for LLVM, please use '3.2', '3.3' or '3.4'")
    sys.exit(-1)
  elif compiler_name == 'gcc' and (compiler_version not in gcc_versions):
    print("invalid version given for GCC, please use '4.6.4' or '4.8.1'")
    sys.exit(-1)

def writeSheetHeader(wsheet, armthumb):
  global header_col, header_row, nr_opts, opts

  col = header_col
  if armthumb == 'Thumb':
    col += nr_opts*2

  # ARM/Thumb
  wsheet.merge_cells(start_row = header_row  , start_column = col, end_row = header_row  , end_column = col+nr_opts*2-1)
  wsheet.cell(row = header_row, column = col).value = armthumb
  wsheet.cell(row = header_row, column = col).style = centerStyle
  # Dynamic
  wsheet.merge_cells(start_row = header_row+1, start_column = col, end_row = header_row+1, end_column = col+nr_opts-1)
  wsheet.cell(row = header_row+1, column = col).value = 'dynamic'
  wsheet.cell(row = header_row+1, column = col).style = centerStyle
  # Static
  wsheet.merge_cells(start_row = header_row+1, start_column = col+nr_opts, end_row = header_row+1, end_column = col+nr_opts*2-1)
  wsheet.cell(row = header_row+1, column = col+nr_opts).value = 'static'
  wsheet.cell(row = header_row+1, column = col+nr_opts).style = centerStyle

  # Compiler optimization levels
  tmp_col = col
  for optlevel in (opts):
    # ... dynamic
    wsheet.cell(row = header_row+2, column = tmp_col).value = optlevel
    wsheet.cell(row = header_row+2, column = tmp_col).style = centerStyle
    # ... static
    wsheet.cell(row = header_row+2, column = tmp_col+nr_opts).value = optlevel
    wsheet.cell(row = header_row+2, column = tmp_col+nr_opts).style = centerStyle

    tmp_col += 1

def benchmarkIndex(benchname):
  global benchmarks

  counter = 0
  for x in benchmarks:
    if benchname in x[0]:
      return counter

    counter += 1

def writeBenchmarkResults(wsheet, benchmark, col, codegain, totalgain, style):
  global benchname_row

  # 2 rows per benchmark
  row = benchname_row + benchmarkIndex(benchmark)*2

  wsheet.cell(row = row  , column = col).value = codegain
  wsheet.cell(row = row  , column = col).style = style
  wsheet.cell(row = row+1, column = col).value = totalgain
  wsheet.cell(row = row+1, column = col).style = style

def writeAllBenchmarkResults(wsheet, name, style):
  global header_col, nr_opts, nr_linkages, arch_mul

  cur_col = header_col

  for j in range(0, nr_opts*nr_linkages*arch_mul):
    writeBenchmarkResults(wsheet, name, cur_col, 'N/A', 'N/A', style)
    cur_col += 1

def isValidBenchmark(name):
  global benchamarks, is_llvm, is_android

  name, ok_llvm, ok_android, ok_diablo = benchmarks[benchmarkIndex(name)]
  return not((not(ok_llvm) and is_llvm) or (not(ok_android) and is_android) or not(ok_diablo))

def writeBenchmarkNames(wsheet):
  global benchname_col, benchname_row

  cur_row = benchname_row

  for name, ok_llvm, ok_android, ok_diablo in (benchmarks):
    wsheet.merge_cells(start_row = cur_row, start_column = benchname_col, end_row = cur_row+1, end_column = benchname_col)
    wsheet.cell(row = cur_row, column = benchname_col).value = name
    wsheet.cell(row = cur_row, column = benchname_col).style = Style(alignment=Alignment(horizontal='left', vertical='center'))

    naStyle = Style(font=Font(size=8), alignment=Alignment(horizontal='center', vertical='center'))

    if not(isValidBenchmark(name)):
      writeAllBenchmarkResults(wsheet, name, naStyle)

    cur_row += 2

def writeSheetHeader(wsheet):
  global header_col, header_row, nr_opts, nr_linkages, arch_mul, compiler_name, compiler_version, centerStyle

  wsheet.merge_cells(start_row = header_row, start_column = header_col, end_row = header_row, end_column = header_col+nr_opts*nr_linkages*arch_mul-1)
  wsheet.cell(row = header_row, column = header_col).value = compiler_name+' '+compiler_version
  wsheet.cell(row = header_row, column = header_col).style = centerStyle
  header_row += 1

  tmp_col = header_col

  variants = ['ARM', 'Thumb']
  if arch_name == 'i486':
    variants = ['']

  for i in variants:
    # header cell for variant
    wsheet.merge_cells(start_row = header_row, start_column = tmp_col, end_row = header_row, end_column = tmp_col+nr_opts*nr_linkages-1)
    wsheet.cell(row = header_row, column = tmp_col).value = i
    wsheet.cell(row = header_row, column = tmp_col).style = centerStyle

    for j in ['static', 'dynamic', 'PIE']:
      wsheet.merge_cells(start_row = header_row+1, start_column = tmp_col, end_row = header_row+1, end_column = tmp_col+nr_opts-1)
      wsheet.cell(row = header_row+1, column = tmp_col).value = j
      wsheet.cell(row = header_row+1, column = tmp_col).style = centerStyle

      for opt in opts:
        wsheet.cell(row = header_row+2, column = tmp_col).value = opt
        wsheet.cell(row = header_row+2, column = tmp_col).style = centerStyle
        tmp_col += 1

def skeletonizeSheet(wsheet):
  writeSheetHeader(wsheet)
  writeBenchmarkNames(wsheet)

# parse command-line arguments
parse_args()

# interpret parameters ==========================
if os_name == 'android':
  is_android = True

# default: ARM/Thumb
if arch_name == 'i486':
  arch_mul = 1

# default: gcc
if compiler_name == 'llvm':
  is_llvm = True

if is_llvm:
  header_col += llvm_versions.index(compiler_version)*nr_opts*nr_linkages*arch_mul
else:
  header_col += gcc_versions.index(compiler_version)*nr_opts*nr_linkages*arch_mul

# default: dynamic
is_static = False
is_pie = False
if linkage == 'static':
  is_static = True
elif linkage == 'pie':
  is_static = False
  is_pie = True

# results column
tmp_col = linkage_names.index(linkage)*nr_opts + opts.index(input_opt)
if arch_name == 'thumb':
  tmp_col += nr_opts*nr_linkages
results_col = header_col + tmp_col

# ===============================================

# create a new or open an existing workbook
if os.path.isfile(workbookname):
  wbook = openpyxl.load_workbook(workbookname)
else:
  wbook = openpyxl.Workbook()

# create a new sheet if necessary
sheet_name = os_name+" - "+compiler_name
if sheet_name in wbook.get_sheet_names():
  wsheet = wbook.get_sheet_by_name(sheet_name)
else:
  wsheet = wbook.create_sheet()
  wsheet.title = sheet_name

# fill in the skeleton
skeletonizeSheet(wsheet)

# read input file
with open(input_file) as f:
  for line in f:
    benchname, diablo_exit, status, codegain, totalgain, clockwall, clockcpu, diablocpu, programsys = line.split(',')

    if isValidBenchmark(benchname):
      color = colors.BLACK

      if not(int(diablo_exit) == 0):
        codegain = 'Diablo'
        totalgain = diablo_exit
        color = colors.RED
      elif not(status == 'OK'):
        codegain = status
        totalgain = ''
        color = colors.RED
      
      writeBenchmarkResults(wsheet, benchname, results_col, codegain, totalgain, Style(font = Font(color=color)))

wbook.save(workbookname)
